"""
2026-05-13 통합 엑셀 → import-payload.json 변환.

입력:
  - H:/Data/대솔이엘_관리현장_20260513_통합.xlsx
  - scripts/site-import/site_name_mapping_final.csv  (excel_name, existing_name_candidate, similarity, action, manual_decision)
  - scripts/site-import/site_name_unmatched_existing.csv  (existing_name_with_no_excel_match)
  - src/data/sites.json (id 보존용)
  - src/data/elevators.json (id 보존용)

출력:
  - scripts/site-import/import-payload.json
      {
        sites:    [ { id?, name, alias, ledgerNo, jobNo, companyType, siteKind,
                      contractType, contractDate, contractStart, contractEnd,
                      warrantyEnd, primaryInspector, subInspector, subInspector2,
                      sitePhone, siteMobile, fax, managerEmail, address, note, vendor,
                      emergencyDevices: [{unit, slot, number}] } ],
        elevators:[ { id?, siteName, unitName, elevatorNo, modelName } ],
        sitesToDelete:     [name, ...],
        elevatorsToDelete: [id, ...]
      }
  - scripts/site-import/import-report.txt
"""
from __future__ import annotations
import csv
import json
import os
import sys
from collections import defaultdict
from datetime import datetime
from typing import Any

import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
XLSX = r"H:\Data\대솔이엘_관리현장_20260513_통합.xlsx"
MAP_CSV = os.path.join(ROOT, "scripts", "site-import", "site_name_mapping_final.csv")
UNM_CSV = os.path.join(ROOT, "scripts", "site-import", "site_name_unmatched_existing.csv")
SITES_JSON = os.path.join(ROOT, "src", "data", "sites.json")
ELEVS_JSON = os.path.join(ROOT, "src", "data", "elevators.json")
PAYLOAD = os.path.join(ROOT, "scripts", "site-import", "import-payload.json")
REPORT = os.path.join(ROOT, "scripts", "site-import", "import-report.txt")


def to_iso(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v).strip() or None


def to_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def main() -> int:
    # ── 매핑 로드 ──────────────────────────────────────
    rename_map: dict[str, str] = {}   # existing_name → excel_name
    new_names: set[str] = set()
    for row in csv.DictReader(open(MAP_CSV, encoding="utf-8-sig")):
        dec = row["manual_decision"].strip()
        if dec.startswith("merge:"):
            existing = dec[len("merge:"):]
            rename_map[existing] = row["excel_name"]
        elif dec == "new":
            new_names.add(row["excel_name"])
        elif dec == "skip":
            pass

    delete_names: set[str] = set()
    for row in csv.reader(open(UNM_CSV, encoding="utf-8-sig")):
        if not row or row[0] == "existing_name_with_no_excel_match":
            continue
        delete_names.add(row[0])

    # ── 기존 데이터 로드 (id 보존용) ───────────────────
    existing_sites = json.load(open(SITES_JSON, encoding="utf-8"))
    existing_elevs = json.load(open(ELEVS_JSON, encoding="utf-8"))

    site_id_by_name: dict[str, int] = {}
    for s in existing_sites:
        site_id_by_name[s["name"]] = s["id"]
    # rename target도 새 이름 키로도 등록
    for old_name, new_name in rename_map.items():
        if old_name in site_id_by_name:
            site_id_by_name[new_name] = site_id_by_name[old_name]

    elev_id_by_key: dict[tuple[str, str | None], int] = {}
    for e in existing_elevs:
        key = (e["siteName"], e.get("unitName"))
        elev_id_by_key[key] = e["id"]
    # rename된 site의 호기들도 새 이름 키로 추가
    for old_name, new_name in rename_map.items():
        for e in existing_elevs:
            if e["siteName"] == old_name:
                elev_id_by_key[(new_name, e.get("unitName"))] = e["id"]

    # ── 엑셀 파싱 ───────────────────────────────────
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    def H(n: str) -> int: return header.index(n)
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    ep_idx = [H("비상통화장치")] + [H(f"비상통화장치{i}") for i in range(2, 11)]

    sites_out: dict[str, dict[str, Any]] = {}
    elevs_out: list[dict[str, Any]] = []

    for r in rows:
        name = to_str(r[H("현장명")])
        if not name:
            continue
        # 동일 현장명에 대해 첫 행만 site 필드로 사용 (이후 행은 비상통화장치만 누적)
        if name not in sites_out:
            sites_out[name] = {
                "name":             name,
                "alias":            to_str(r[H("현장별칭")]),
                "ledgerNo":         to_str(r[H("원장번호")]),
                "jobNo":            to_str(r[H("proj_no")]),
                "companyType":      to_str(r[H("회사구분")]),
                "siteKind":         to_str(r[H("현장유형")]),
                "contractType":     to_str(r[H("계약구분")]),
                "contractDate":     to_iso(r[H("계약일자")]),
                "contractStart":    to_iso(r[H("계약시작")]),
                "contractEnd":      to_iso(r[H("계약만료")]),
                "warrantyEnd":      to_iso(r[H("하자기간")]),
                "primaryInspector": to_str(r[H("주점검자")]),
                "subInspector":     to_str(r[H("보조점검자1")]),
                "subInspector2":    to_str(r[H("보조점검자2")]),
                "sitePhone":        to_str(r[H("현장전화")]),
                "siteMobile":       to_str(r[H("현장연락처2")]),
                "fax":              to_str(r[H("팩스")]),
                "managerEmail":     to_str(r[H("담당자 메일")]),
                "address":          to_str(r[H("승강기 소재지")]),
                "note":             to_str(r[H("비고")]),
                "vendor":           to_str(r[H("거래처")]),
                "emergencyDevices": [],
            }
        # 호기 단위로 비상통화장치 1~10 누적 (사용자 결정: 현장 단위 jsonb로만 통합)
        unit = to_str(r[H("호기명")])
        eno = to_str(r[H("승강기번호")])
        model = to_str(r[H("기종명")])

        for slot_i, idx in enumerate(ep_idx, start=1):
            v = to_str(r[idx])
            if v:
                sites_out[name]["emergencyDevices"].append({
                    "unit": unit or "",
                    "slot": slot_i,
                    "number": v,
                })

        elevs_out.append({
            "siteName":   name,
            "unitName":   unit,
            "elevatorNo": eno,
            "modelName":  model,
        })

    # ── id 채우기 ───────────────────────────────────
    payload_sites = []
    for name, s in sites_out.items():
        sid = site_id_by_name.get(name)
        if sid:
            s["id"] = sid
        payload_sites.append(s)

    payload_elevators = []
    for e in elevs_out:
        eid = elev_id_by_key.get((e["siteName"], e["unitName"]))
        if eid:
            e["id"] = eid
        payload_elevators.append(e)

    # ── 삭제 후보 ───────────────────────────────────
    sites_to_delete = sorted(delete_names)
    # 삭제 대상 sites의 elevators ids
    elevators_to_delete = [
        e["id"] for e in existing_elevs if e["siteName"] in delete_names
    ]
    # 매핑된 기존 사이트의 unit_name=null placeholder 호기 → 정리 대상
    placeholder_elevator_ids = []
    for old_name, new_name in rename_map.items():
        for e in existing_elevs:
            if e["siteName"] == old_name and not e.get("unitName"):
                placeholder_elevator_ids.append(e["id"])

    # ── 출력 ───────────────────────────────────────
    payload = {
        "sites":                 payload_sites,
        "elevators":             payload_elevators,
        "renameMap":             rename_map,
        "sitesToDelete":         sites_to_delete,
        "elevatorsToDelete":     elevators_to_delete,
        "placeholderElevatorIds": placeholder_elevator_ids,
    }
    with open(PAYLOAD, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    # ── 리포트 ──────────────────────────────────────
    new_sites = [s for s in payload_sites if "id" not in s]
    upd_sites = [s for s in payload_sites if "id" in s]
    new_elevs = [e for e in payload_elevators if "id" not in e]
    upd_elevs = [e for e in payload_elevators if "id" in e]
    lines = [
        "=== Import Payload Report ===",
        f"xlsx                    : {XLSX}",
        f"payload                 : {PAYLOAD}",
        "",
        f"sites total             : {len(payload_sites)}",
        f"  new                   : {len(new_sites)}",
        f"  update (id preserved) : {len(upd_sites)}",
        "",
        f"elevators total         : {len(payload_elevators)}",
        f"  new                   : {len(new_elevs)}",
        f"  update (id preserved) : {len(upd_elevs)}",
        "",
        f"renameMap               : {len(rename_map)}",
        *[f"  {old} -> {new}" for old, new in rename_map.items()],
        "",
        f"sites to delete         : {len(sites_to_delete)}",
        *[f"  - {n}" for n in sites_to_delete],
        f"elevators to delete (orphan): {len(elevators_to_delete)}",
        f"placeholder elevators to clean : {len(placeholder_elevator_ids)}",
    ]
    with open(REPORT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print("\n".join(lines))
    return 0


if __name__ == "__main__":
    sys.exit(main())
